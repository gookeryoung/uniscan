"""扫描报告导出单元测试。

覆盖 ``fuscan.scanner.export`` 模块的 PDF/Excel 二进制导出与 ``save_report``
按扩展名分发的文件保存逻辑。测试用例从 ``test_scanner.py`` 迁移而来，
原 ``ScanReport.to_pdf``/``to_excel``/``save_report`` 方法已拆分到本模块。
"""

from __future__ import annotations

from pathlib import Path

from fuscan.rules.model import Severity
from fuscan.scanner import ScanReport, ScanResult
from fuscan.scanner.export import export_excel, export_pdf, save_report
from fuscan.scanner.result import RuleHit, ScanStats


def _build_report(tmp_path: Path) -> ScanReport:
    """构造测试报告：3 个文件命中 2 条规则，分属 WARNING/CRITICAL 两个等级。"""
    results = (
        ScanResult(
            path=tmp_path / "secret.txt" / "a.txt",
            size=10,
            hits=(
                RuleHit("敏感文件名", Severity.WARNING, "d1", match_count=1),
                RuleHit("密钥内容", Severity.CRITICAL, "d2", match_count=2),
            ),
        ),
        ScanResult(
            path=tmp_path / "secret.txt" / "b.txt",
            size=20,
            hits=(RuleHit("密钥内容", Severity.CRITICAL, "d3", match_count=3),),
        ),
        ScanResult(path=tmp_path / "clean.txt", size=0, hits=()),
    )
    stats = ScanStats(
        total_files=3,
        scanned_files=3,
        matched_files=2,
        skipped_files=0,
        errors=0,
        duration_seconds=0.5,
        total_matches=6,
    )
    return ScanReport(root=tmp_path, results=results, stats=stats)


class TestExportPdf:
    def test_export_pdf_returns_bytes_with_header(self, tmp_path: Path) -> None:
        """export_pdf 应返回 PDF 二进制数据，以 %PDF- 开头。"""
        report = _build_report(tmp_path)
        data = export_pdf(report)
        assert isinstance(data, bytes)
        assert data[:5] == b"%PDF-"

    def test_export_pdf_empty_hits(self, tmp_path: Path) -> None:
        """空命中报告也能生成 PDF，仍以 %PDF- 开头。"""
        report = ScanReport(root=tmp_path, results=(), stats=ScanStats())
        data = export_pdf(report)
        assert data[:5] == b"%PDF-"

    def test_export_pdf_contains_keywords(self, tmp_path: Path) -> None:
        """PDF 文本流应包含 PDF 结构标记（CID 字体下中文不可直接 grep）。"""
        report = _build_report(tmp_path)
        data = export_pdf(report)
        # PDF 中文字以 CID 编码无法直接 grep，但 PDF 结构标记应可见
        assert b"/Type /Catalog" in data or b"/Pages" in data


class TestExportExcel:
    def test_export_excel_returns_zip_archive(self, tmp_path: Path) -> None:
        """export_excel 应返回 xlsx 二进制数据（zip 格式，PK 开头）。"""
        report = _build_report(tmp_path)
        data = export_excel(report)
        assert isinstance(data, bytes)
        # xlsx 是 zip 压缩包，开头为 PK\x03\x04
        assert data[:2] == b"PK"

    def test_export_excel_empty_hits(self, tmp_path: Path) -> None:
        """空命中报告也能生成 Excel。"""
        report = ScanReport(root=tmp_path, results=(), stats=ScanStats())
        data = export_excel(report)
        assert data[:2] == b"PK"

    def test_export_excel_roundtrip(self, tmp_path: Path) -> None:
        """生成的 xlsx 应能被 openpyxl 读回，且工作表名称正确。"""
        from openpyxl import load_workbook

        report = _build_report(tmp_path)
        data = export_excel(report)
        import io as _io

        wb = load_workbook(_io.BytesIO(data))
        assert "扫描汇总" in wb.sheetnames
        assert "命中明细" in wb.sheetnames
        # 命中明细表头应在第 1 行
        headers = [c.value for c in wb["命中明细"][1]]
        assert headers == ["路径", "大小", "严重等级", "规则", "描述", "匹配数", "详情"]


class TestSaveReport:
    def test_save_report_csv(self, tmp_path: Path) -> None:
        """save_report 按 .csv 扩展名写入 UTF-8 文本。"""
        report = _build_report(tmp_path)
        target = tmp_path / "out.csv"
        save_report(report, target)
        content = target.read_text(encoding="utf-8")
        assert content.startswith("path,size,severity,rule,description,match_count,detail")

    def test_save_report_json(self, tmp_path: Path) -> None:
        """save_report 按 .json 扩展名写入 JSON 文本。"""
        import json as _json

        report = _build_report(tmp_path)
        target = tmp_path / "out.json"
        save_report(report, target)
        data = _json.loads(target.read_text(encoding="utf-8"))
        assert data["root"] == str(tmp_path)

    def test_save_report_txt_fallback(self, tmp_path: Path) -> None:
        """save_report 对 .txt 扩展名按 text 格式写入。"""
        report = _build_report(tmp_path)
        target = tmp_path / "out.txt"
        save_report(report, target)
        content = target.read_text(encoding="utf-8")
        assert "扫描路径:" in content

    def test_save_report_pdf(self, tmp_path: Path) -> None:
        """save_report 按 .pdf 扩展名写入二进制。"""
        report = _build_report(tmp_path)
        target = tmp_path / "out.pdf"
        save_report(report, target)
        data = target.read_bytes()
        assert data[:5] == b"%PDF-"

    def test_save_report_xlsx(self, tmp_path: Path) -> None:
        """save_report 按 .xlsx 扩展名写入二进制。"""
        report = _build_report(tmp_path)
        target = tmp_path / "out.xlsx"
        save_report(report, target)
        data = target.read_bytes()
        assert data[:2] == b"PK"

    def test_save_report_unknown_ext_falls_back_to_text(self, tmp_path: Path) -> None:
        """save_report 对未知扩展名（非 csv/json/pdf/xlsx）按 text 写入。"""
        report = _build_report(tmp_path)
        target = tmp_path / "out.log"
        save_report(report, target)
        content = target.read_text(encoding="utf-8")
        assert "扫描路径:" in content
