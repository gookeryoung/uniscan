"""电子表格提取器：XLSX 与 ODS。

XLSX 使用 openpyxl 提取所有工作表单元格文本。
ODS 使用 odfpy 提取表格内容。
"""

from __future__ import annotations

import logging
from pathlib import Path

from fuscan.extractors.base import Extractor, ExtractorError

__all__ = ["OdsExtractor", "XlsxExtractor"]

logger = logging.getLogger(__name__)

_MAX_ROWS = 10000
_MAX_COLS = 256


class XlsxExtractor(Extractor):
    """XLSX 电子表格文本提取器。"""

    def __init__(self, max_rows: int = _MAX_ROWS, max_cols: int = _MAX_COLS) -> None:
        self._max_rows = max_rows
        self._max_cols = max_cols

    @property
    def supported_extensions(self) -> tuple[str, ...]:
        return ("xlsx", "xlsm")

    def extract(self, path: Path) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ExtractorError("openpyxl 未安装，无法提取 XLSX") from exc

        try:
            wb = load_workbook(str(path), read_only=True, data_only=True)
        except Exception as exc:
            raise ExtractorError(f"XLSX 解析失败: {path}: {exc}") from exc

        parts: list[str] = []
        try:
            for sheet in wb:
                sheet_texts = self._extract_sheet(sheet)
                if sheet_texts:
                    parts.append(f"--- 工作表: {sheet.title} ---")
                    parts.extend(sheet_texts)
        finally:
            wb.close()

        return "\n".join(parts)

    def _extract_sheet(self, sheet: object) -> list[str]:
        """提取单个工作表的文本。"""
        texts: list[str] = []
        for row_count, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if row_count > self._max_rows:
                logger.debug("工作表行数超过上限 %d，截断", self._max_rows)
                break
            cell_texts = []
            for col_idx, cell in enumerate(row):
                if col_idx >= self._max_cols:
                    break
                if cell is not None:
                    cell_str = str(cell).strip()
                    if cell_str:
                        cell_texts.append(cell_str)
            if cell_texts:
                texts.append("\t".join(cell_texts))
        return texts


class OdsExtractor(Extractor):
    """ODS 电子表格文本提取器（OpenDocument Spreadsheet）。"""

    @property
    def supported_extensions(self) -> tuple[str, ...]:
        return ("ods",)

    def extract(self, path: Path) -> str:
        try:
            from odf.opendocument import load
            from odf.table import TableCell, TableRow
        except ImportError as exc:
            raise ExtractorError("odfpy 未安装，无法提取 ODS") from exc

        try:
            doc = load(str(path))
        except Exception as exc:
            raise ExtractorError(f"ODS 解析失败: {path}: {exc}") from exc

        parts: list[str] = []
        for row in doc.getElementsByType(TableRow):
            row_texts = []
            for cell in row.getElementsByType(TableCell):
                text = self._extract_cell_text(cell)
                if text:
                    row_texts.append(text)
            if row_texts:
                parts.append("\t".join(row_texts))

        return "\n".join(parts)

    def _extract_cell_text(self, cell: object) -> str:
        """提取单元格内所有文本节点。"""
        try:
            return str(cell).strip()
        except Exception:
            return ""
