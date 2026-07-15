"""扫描结果数据结构。

除 ``MatchResult``/``ProgressInfo`` 外，``ScanResult``/``ScanStats``/``ScanReport``
均提供数据层方法（``rule_names``/``filter``/``group_by_*``/``to_*``/``summary``/
``file_info_html``/``notification_message``/``to_format``），将"如何序列化、如何
筛选、如何分组、如何格式化展示文本"下沉到 dataclass，CLI/GUI 仅做展示，
避免展示层重复实现相同逻辑。

模块级 ``format_size`` 为字节数人类可读格式化，原属 GUI 层，提升至数据层后
供 dataclass 与 GUI 共享。
"""

from __future__ import annotations

import contextlib
import csv
import datetime
import html
import io
import json
from dataclasses import asdict, dataclass, field
from pathlib import Path

from fuscan.rules.model import Severity

__all__ = [
    "MatchResult",
    "ProgressInfo",
    "RuleHit",
    "ScanReport",
    "ScanResult",
    "ScanStats",
    "format_size",
]


def format_size(size: int) -> str:
    """将字节数格式化为人类可读字符串（B/KB/MB/GB）。"""
    if size < 1024:
        return f"{size} B"
    if size < 1024 * 1024:
        return f"{size / 1024:.1f} KB"
    if size < 1024 * 1024 * 1024:
        return f"{size / (1024 * 1024):.1f} MB"
    return f"{size / (1024 * 1024 * 1024):.2f} GB"


@dataclass(frozen=True)
class MatchResult:
    """单次匹配求值结果。

    ``match_text`` 存储匹配到的原始文本（regex 模式为首个 ``m.group(0)``，
    其他模式为 ``pattern``），供 GUI 高亮定位使用，避免 ``repr`` 转义导致的失真。
    ``match_count`` 为该次求值实际匹配到的文本条数（同一规则在同一文件中
    可能匹配多处，如多处密码、多处密钥），用于区分"命中规则数"与"匹配条数"。
    ``target`` 为匹配目标类型（"filename"/"content"/"path"），叶子匹配器设置，
    组合规则为空字符串。GUI 根据 ``target`` 判断是否在内容预览中搜索高亮位置。
    ``match_texts`` 为该次求值命中的所有文本（去重保序，含组合规则子匹配文本），
    用于 AND/OR 等组合规则标记每个命中的内容，避免仅标记单一匹配规则的内容。
    ``match_description`` 为该匹配项的可选描述（来自 MatchSpec.description），
    便于用户理解匹配规则含义，在 GUI 详情表与导出结果中展示。
    """

    matched: bool
    detail: str = ""
    match_text: str = ""
    match_count: int = 1
    target: str = ""
    match_texts: tuple[str, ...] = ()
    match_description: str = ""


@dataclass(frozen=True)
class ProgressInfo:
    """扫描进度信息（实时反馈给 UI）。"""

    current_file: str = ""
    scanned: int = 0
    total: int = 0
    skipped: int = 0
    matched: int = 0
    errors: int = 0
    elapsed: float = 0.0
    # 匹配文本条数（同一规则在同一文件的多处匹配分别计数）
    matches: int = 0
    # 跳过的目录路径（最近 500 条，避免无限增长）
    skipped_dirs: tuple[str, ...] = ()
    # 命中的 (文件路径, 规则名) 列表（最近 500 条）
    matched_files: tuple[tuple[str, str], ...] = ()

    def summary(self) -> str:
        """返回实时进度状态栏文本（含速度计算）。

        与 ``ScanStats.summary`` 不同，此处不含 ``total_files`` 与 ``cancelled``
        前缀，仅展示当前已扫描的实时指标，供 GUI 进度回调直接调用。
        """
        speed = self.scanned / self.elapsed if self.elapsed > 0 else 0.0
        return (
            f"已扫描 {self.scanned} | 跳过 {self.skipped} | "
            f"命中 {self.matched} | 条数 {self.matches} | 错误 {self.errors} | "
            f"已用 {self.elapsed:.1f}s | 速度 {speed:.0f} 文件/s"
        )


@dataclass(frozen=True)
class RuleHit:
    """规则命中记录：一条规则对一个文件的命中信息。

    ``match_text`` 为匹配到的原始文本，供 GUI 高亮定位使用；
    对于组合规则（and/or/not）无单一匹配文本时为空字符串。
    ``match_count`` 为该规则在该文件实际匹配到的文本条数（如多处密码各算 1 条），
    用于区分"命中规则数"与"匹配条数"，避免两者不对等时产生歧义。
    ``target`` 为匹配目标类型（"filename"/"content"/"path"），叶子匹配器设置，
    组合规则为空字符串。GUI 据 ``target=="filename"`` 判断是否在内容预览中
    搜索高亮位置——文件名匹配不应在内容中搜索高亮，否则可能产生误导。
    ``match_texts`` 为该规则命中的所有文本（去重保序，含组合规则子匹配文本），
    用于 AND/OR 等组合规则标记每个命中的内容。
    ``match_description`` 为该匹配项的可选描述（来自 MatchSpec.description），
    便于用户理解匹配规则含义，在 GUI 详情表与导出结果中展示。
    """

    rule_name: str
    severity: Severity
    detail: str
    match_text: str = ""
    match_count: int = 1
    target: str = ""
    match_texts: tuple[str, ...] = ()
    match_description: str = ""


@dataclass(frozen=True)
class ScanResult:
    """单个文件的扫描结果。"""

    path: Path
    size: int
    hits: tuple[RuleHit, ...] = field(default_factory=tuple)
    errors: int = 0

    @property
    def has_hit(self) -> bool:
        return bool(self.hits)

    @property
    def has_error(self) -> bool:
        return self.errors > 0

    @property
    def max_severity(self) -> Severity:
        """该文件命中规则中的最高严重等级。"""
        if not self.hits:
            return Severity.INFO
        order = {Severity.INFO: 0, Severity.WARNING: 1, Severity.CRITICAL: 2}
        return max(self.hits, key=lambda h: order[h.severity]).severity

    @property
    def total_match_count(self) -> int:
        """该文件所有命中规则的匹配文本条数之和。"""
        return sum(h.match_count for h in self.hits)

    @property
    def rule_names(self) -> tuple[str, ...]:
        """该文件命中的规则名（按出现顺序去重）。"""
        seen: set[str] = set()
        names: list[str] = []
        for h in self.hits:
            if h.rule_name not in seen:
                seen.add(h.rule_name)
                names.append(h.rule_name)
        return tuple(names)

    def summary(self) -> str:
        """返回简洁摘要：``N 条规则 / M 处匹配``。"""
        return f"{len(self.hits)} 条规则 / {self.total_match_count} 处匹配"

    def file_info_html(self, extra: str = "") -> str:
        """返回文件元信息 HTML 片段（供 GUI 详情区/对话框共用）。

        含文件路径、大小、修改时间、命中规则数、匹配条数。
        ``extra`` 用于 GUI 追加自身状态相关的字段（如"可切换位置"数），
        为已格式化的 HTML 片段，将以 `` | `` 分隔附加在末尾。

        修改时间通过 ``Path.stat`` 获取，失败时显示"无法获取"。
        """
        try:
            mtime = datetime.datetime.fromtimestamp(self.path.stat().st_mtime)
            mtime_str = mtime.strftime("%Y-%m-%d %H:%M:%S")
        except OSError:
            mtime_str = "无法获取"

        info = (
            f"<b>文件路径:</b> {html.escape(str(self.path))}<br>"
            f"<b>文件大小:</b> {format_size(self.size)} ({self.size} 字节)<br>"
            f"<b>修改时间:</b> {html.escape(mtime_str)}<br>"
            f"<b>命中规则数:</b> {len(self.hits)} | <b>匹配条数:</b> {self.total_match_count}"
        )
        if extra:
            info += f" | {extra}"
        return info


@dataclass(frozen=True)
class ScanStats:
    """扫描统计。"""

    total_files: int = 0
    scanned_files: int = 0
    matched_files: int = 0
    skipped_files: int = 0
    errors: int = 0
    duration_seconds: float = 0.0
    # 所有命中规则的匹配文本条数总和（区别于 matched_files 的命中文件数）
    total_matches: int = 0

    def summary(self, *, cancelled: bool = False) -> str:
        """返回状态栏摘要文本。

        :param cancelled: 是否在摘要前缀"已取消"，GUI/CLI 取消场景共用。
        """
        prefix = "已取消" if cancelled else "完成"
        return (
            f"{prefix}: 总计 {self.total_files} | 扫描 {self.scanned_files} | "
            f"跳过 {self.skipped_files} | 命中 {self.matched_files} | "
            f"条数 {self.total_matches} | 错误 {self.errors} | "
            f"耗时 {self.duration_seconds:.2f}s"
        )


@dataclass(frozen=True)
class ScanReport:
    """完整扫描报告。

    提供数据层操作（``filter``/``group_by_*``/``to_*``/``rule_names``/``summary``），
    将序列化、筛选、分组逻辑下沉到 dataclass，CLI/GUI 仅做展示，
    避免展示层重复实现相同逻辑。
    """

    root: Path
    results: tuple[ScanResult, ...] = field(default_factory=tuple)
    stats: ScanStats = field(default_factory=ScanStats)
    cancelled: bool = False

    @property
    def hits(self) -> tuple[ScanResult, ...]:
        """仅返回有命中的结果。"""
        return tuple(r for r in self.results if r.has_hit)

    @property
    def rule_names(self) -> tuple[str, ...]:
        """所有命中结果涉及的规则名（按首次出现顺序去重）。"""
        seen: set[str] = set()
        names: list[str] = []
        for r in self.hits:
            for name in r.rule_names:
                if name not in seen:
                    seen.add(name)
                    names.append(name)
        return tuple(names)

    def summary(self) -> str:
        """返回状态栏摘要文本（自动识别 ``cancelled`` 标志）。"""
        return self.stats.summary(cancelled=self.cancelled)

    def notification_message(self) -> str:
        """返回托盘/通知消息正文。

        形如 ``发现 3 个文件命中规则，共 7 处匹配``；无命中时返回 ``未发现命中``。
        """
        if not self.hits:
            return "未发现命中"
        return f"发现 {len(self.hits)} 个文件命中规则，共 {self.stats.total_matches} 处匹配"

    def to_format(self, fmt: str) -> str:
        """按格式名渲染报告（``json``/``csv``/``text``）。

        调用方无需维护 if-else 调度，未知格式回退到 ``text``。
        """
        if fmt == "json":
            return self.to_json()
        if fmt == "csv":
            return self.to_csv()
        return self.to_text()

    def filter(self, path_query: str = "", rule_name: str = "") -> ScanReport:
        """按路径子串与规则名筛选，返回新的 ScanReport（不修改原对象）。

        - ``path_query``：大小写不敏感的路径子串，空字符串跳过路径过滤
        - ``rule_name``：规则名精确匹配；非空时仅保留该规则命中，
          且每个 ScanResult 的 hits 被过滤为仅该规则的命中

        stats 不变（仍代表整体扫描统计），仅 results 被过滤。
        """
        query = path_query.strip().lower()
        if not query and not rule_name:
            return self
        filtered: list[ScanResult] = []
        for sr in self.hits:
            if query and query not in str(sr.path).lower():
                continue
            if rule_name:
                matching_hits = tuple(h for h in sr.hits if h.rule_name == rule_name)
                if not matching_hits:
                    continue
                filtered.append(ScanResult(path=sr.path, size=sr.size, hits=matching_hits, errors=sr.errors))
            else:
                filtered.append(sr)
        return ScanReport(
            root=self.root,
            results=tuple(filtered),
            stats=self.stats,
            cancelled=self.cancelled,
        )

    def group_by_rule(self) -> dict[str, list[tuple[ScanResult, RuleHit]]]:
        """按规则名分组：``{规则名: [(ScanResult, RuleHit), ...]}``。

        同一文件若被同一规则多次命中（理论上不会，但保留兼容）会重复出现；
        同一规则在不同文件的命中分别作为列表项。
        """
        groups: dict[str, list[tuple[ScanResult, RuleHit]]] = {}
        for sr in self.hits:
            for hit in sr.hits:
                groups.setdefault(hit.rule_name, []).append((sr, hit))
        return groups

    def group_by_severity(self) -> dict[Severity, list[ScanResult]]:
        """按文件最高严重等级分组：``{Severity: [ScanResult, ...]}``。"""
        groups: dict[Severity, list[ScanResult]] = {}
        for sr in self.hits:
            groups.setdefault(sr.max_severity, []).append(sr)
        return groups

    def to_json(self) -> str:
        """将扫描报告转换为 JSON 字符串。"""
        data = {
            "root": str(self.root),
            "stats": asdict(self.stats),
            "cancelled": self.cancelled,
            "hits": [
                {
                    "path": str(r.path),
                    "size": r.size,
                    "max_severity": r.max_severity.value,
                    "match_count": r.total_match_count,
                    "rules": [asdict(h) for h in r.hits],
                }
                for r in self.hits
            ],
        }
        return json.dumps(data, ensure_ascii=False, indent=2)

    def to_csv(self) -> str:
        """将扫描报告转换为 CSV 字符串（每行一条规则命中）。"""
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["path", "size", "severity", "rule", "description", "match_count", "detail"])
        for r in self.hits:
            for hit in r.hits:
                writer.writerow(
                    [
                        str(r.path),
                        r.size,
                        hit.severity.value,
                        hit.rule_name,
                        hit.match_description,
                        hit.match_count,
                        hit.detail,
                    ]
                )
        return buf.getvalue()

    def to_text(self) -> str:
        """将扫描报告转换为可读文本（含路径、统计与逐文件命中详情）。"""
        lines: list[str] = []
        lines.append(f"扫描路径: {self.root}")
        lines.append(f"统计: {self.stats.summary(cancelled=self.cancelled)}")
        lines.append("")
        if not self.hits:
            lines.append("未发现命中项。")
            return "\n".join(lines)
        lines.append(f"命中项 ({len(self.hits)}):")
        for result in self.hits:
            try:
                rel = result.path.relative_to(self.root)
            except ValueError:
                rel = result.path
            lines.append(f"  {rel} (规则 {len(result.hits)} / 条数 {result.total_match_count})")
            for hit in result.hits:
                # 描述非空时附加在规则名后，便于用户理解匹配规则含义
                rule_label = f"{hit.rule_name} - {hit.match_description}" if hit.match_description else hit.rule_name
                lines.append(f"    [{hit.severity.value}] {rule_label} (条数 {hit.match_count}): {hit.detail}")
        return "\n".join(lines)

    def to_pdf(self) -> bytes:
        """将扫描报告转换为 PDF 二进制数据。

        使用 reportlab 生成，含标题、扫描统计、命中文件表格。
        中文字体使用 ``STSong-Light`` CID 字体（跨平台一致，无需字体文件）。
        """
        return _build_pdf(self)

    def to_excel(self) -> bytes:
        """将扫描报告转换为 Excel 二进制数据（.xlsx）。

        使用 openpyxl 生成，含两个工作表：

        - "扫描汇总"：扫描路径、时间、统计信息
        - "命中明细"：每行一条规则命中（含描述列）
        """
        return _build_excel(self)

    def save_report(self, path: Path) -> None:
        """将扫描报告保存到文件，根据扩展名自动选择格式。

        支持的扩展名：``.csv``/``.json``/``.txt``/``.pdf``/``.xlsx``。
        其他扩展名按文本格式输出。
        """
        ext = path.suffix.lower()
        if ext == ".pdf":
            path.write_bytes(self.to_pdf())
            return
        if ext == ".xlsx":
            path.write_bytes(self.to_excel())
            return
        content = self.to_format(ext.lstrip(".") if ext.lstrip(".") in ("csv", "json") else "text")
        path.write_text(content, encoding="utf-8")


def _build_pdf(report: ScanReport) -> bytes:
    """使用 reportlab 构建 PDF 报告。

    结构：封面标题 → 扫描统计 → 命中文件表格。
    """
    # 惰性导入：reportlab 是核心依赖但导入较重
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    # 注册中文字体（跨平台一致，无需字体文件；已注册时忽略）
    font_name = "STSong-Light"
    with contextlib.suppress(Exception):
        pdfmetrics.registerFont(UnicodeCIDFont(font_name))

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4, leftMargin=20 * mm, rightMargin=20 * mm, topMargin=20 * mm, bottomMargin=18 * mm
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("ChTitle", parent=styles["Title"], fontName=font_name, fontSize=20)
    heading_style = ParagraphStyle("ChHeading", parent=styles["Heading2"], fontName=font_name, fontSize=14)
    normal_style = ParagraphStyle("ChNormal", parent=styles["Normal"], fontName=font_name, fontSize=10, leading=16)
    cell_style = ParagraphStyle("ChCell", parent=styles["Normal"], fontName=font_name, fontSize=9, leading=13)

    # reportlab 的 Flowable 是动态类型，story 元素均为其子类（Paragraph/Spacer/Table）
    from typing import Any

    story: list[Any] = []
    story.append(Paragraph("fuscan 扫描报告", title_style))
    story.append(Spacer(1, 8 * mm))

    # 统计信息
    story.append(Paragraph("扫描统计", heading_style))
    story.append(Spacer(1, 3 * mm))
    stat_lines = [
        f"扫描路径: {report.root}",
        f"扫描时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"总文件数: {report.stats.total_files}",
        f"已扫描: {report.stats.scanned_files}",
        f"命中文件: {report.stats.matched_files}",
        f"总匹配数: {report.stats.total_matches}",
        f"跳过: {report.stats.skipped_files}",
        f"错误: {report.stats.errors}",
        f"耗时: {report.stats.duration_seconds:.2f}s",
    ]
    if report.cancelled:
        stat_lines.append("状态: 已取消")
    for line in stat_lines:
        story.append(Paragraph(line, normal_style))
    story.append(Spacer(1, 6 * mm))

    # 命中文件表格
    if not report.hits:
        story.append(Paragraph("未发现命中项。", normal_style))
    else:
        story.append(Paragraph(f"命中文件 ({len(report.hits)})", heading_style))
        story.append(Spacer(1, 3 * mm))
        # 表头
        header = [Paragraph(text, cell_style) for text in ["路径", "大小", "等级", "规则", "描述", "匹配数", "详情"]]
        # 表格行含 Paragraph 与字符串混合，用 list[Any] 容纳
        rows: list[list[Any]] = [header]
        for result in report.hits:
            try:
                rel = str(result.path.relative_to(report.root))
            except ValueError:
                rel = str(result.path)
            for hit in result.hits:
                rows.append(
                    [
                        Paragraph(rel, cell_style),
                        Paragraph(format_size(result.size), cell_style),
                        Paragraph(hit.severity.value, cell_style),
                        Paragraph(hit.rule_name, cell_style),
                        Paragraph(hit.match_description or "", cell_style),
                        Paragraph(str(hit.match_count), cell_style),
                        Paragraph(hit.detail, cell_style),
                    ]
                )
        col_widths = [50 * mm, 18 * mm, 14 * mm, 24 * mm, 24 * mm, 12 * mm, 32 * mm]
        table = Table(rows, colWidths=col_widths, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0887A0")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1DDE2")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F8FA")]),
                ]
            )
        )
        story.append(table)

    doc.build(story)
    return buf.getvalue()


def _build_excel(report: ScanReport) -> bytes:
    """使用 openpyxl 构建 Excel 报告。

    Sheet1 "扫描汇总"：扫描路径、时间、统计信息。
    Sheet2 "命中明细"：每行一条规则命中，表头加粗，严重等级着色。
    """
    # 惰性导入：openpyxl 是核心依赖但导入较重
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Sheet1: 扫描汇总
    ws1 = wb.active
    assert ws1 is not None  # Workbook 新建时 active 一定非空，缩窄 pyrefly 类型推断
    ws1.title = "扫描汇总"
    bold_font = Font(bold=True)
    ws1["A1"] = "fuscan 扫描报告"
    ws1["A1"].font = Font(bold=True, size=14)
    summary_rows = [
        ("扫描路径", str(report.root)),
        ("扫描时间", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("总文件数", report.stats.total_files),
        ("已扫描", report.stats.scanned_files),
        ("命中文件", report.stats.matched_files),
        ("总匹配数", report.stats.total_matches),
        ("跳过", report.stats.skipped_files),
        ("错误", report.stats.errors),
        ("耗时(秒)", round(report.stats.duration_seconds, 2)),
        ("状态", "已取消" if report.cancelled else "完成"),
    ]
    for row_idx, (label, value) in enumerate(summary_rows, start=3):
        ws1.cell(row=row_idx, column=1, value=label).font = bold_font
        ws1.cell(row=row_idx, column=2, value=value)
    ws1.column_dimensions["A"].width = 16
    ws1.column_dimensions["B"].width = 50

    # Sheet2: 命中明细
    ws2 = wb.create_sheet("命中明细")
    headers = ["路径", "大小", "严重等级", "规则", "描述", "匹配数", "详情"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = bold_font
        cell.fill = PatternFill(start_color="0887A0", end_color="0887A0", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    row_idx = 2
    # 严重等级颜色映射
    severity_fills = {
        Severity.CRITICAL: PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid"),
        Severity.WARNING: PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid"),
        Severity.INFO: PatternFill(start_color="D4EFDF", end_color="D4EFDF", fill_type="solid"),
    }
    wrap_align = Alignment(wrap_text=True, vertical="top")
    for result in report.hits:
        try:
            rel = str(result.path.relative_to(report.root))
        except ValueError:
            rel = str(result.path)
        for hit in result.hits:
            ws2.cell(row=row_idx, column=1, value=rel).alignment = wrap_align
            ws2.cell(row=row_idx, column=2, value=format_size(result.size))
            severity_cell = ws2.cell(row=row_idx, column=3, value=hit.severity.value)
            severity_cell.fill = severity_fills.get(hit.severity, PatternFill())
            ws2.cell(row=row_idx, column=4, value=hit.rule_name).alignment = wrap_align
            ws2.cell(row=row_idx, column=5, value=hit.match_description).alignment = wrap_align
            ws2.cell(row=row_idx, column=6, value=hit.match_count)
            ws2.cell(row=row_idx, column=7, value=hit.detail).alignment = wrap_align
            row_idx += 1
    # 列宽
    col_widths = [50, 12, 12, 20, 20, 10, 40]
    for col_idx, width in enumerate(col_widths, start=1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = width
    # 冻结表头
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
