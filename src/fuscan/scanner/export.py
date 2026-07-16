"""扫描报告导出：PDF/Excel 二进制格式与文件保存。

将导出逻辑从 ``scanner.result`` 分离，使 ``ScanReport`` 仅承担数据结构与
数据层方法（查询/筛选/格式化/文本序列化），二进制导出（依赖 reportlab/
openpyxl）集中在本模块，职责单一。

公共 API：

- :func:`export_pdf`：生成 PDF 二进制（reportlab，STSong-Light 中文字体）
- :func:`export_excel`：生成 Excel 二进制（openpyxl，双工作表 + 严重等级着色）
- :func:`save_report`：按文件扩展名自动选择格式写入文件

文本格式（csv/json/text）仍由 :meth:`ScanReport.to_format` 处理，本模块
仅负责二进制格式与文件分发。
"""

from __future__ import annotations

import contextlib
import datetime
import io
from pathlib import Path

from fuscan.rules.model import Severity
from fuscan.scanner.result import ScanReport, format_size

__all__ = ["export_excel", "export_pdf", "save_report"]


def export_pdf(report: ScanReport) -> bytes:
    """将扫描报告转换为 PDF 二进制数据。

    使用 reportlab 生成，含标题、扫描统计、命中文件表格。
    中文字体使用 ``STSong-Light`` CID 字体（跨平台一致，无需字体文件）。
    """
    # 惰性导入：reportlab 是核心依赖但导入较重
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    # 注册中文字体（跨平台一致，无需字体文件；已注册时忽略）
    font_name = "STSong-Light"
    with contextlib.suppress(Exception):
        pdfmetrics.registerFont(UnicodeCIDFont(font_name))

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4, leftMargin=20 * mm, rightMargin=20 * mm, topMargin=20 * mm, bottomMargin=18 * mm
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("ChTitle", parent=styles["Title"], fontName=font_name, fontSize=20)
    heading_style = ParagraphStyle("ChHeading", parent=styles["Heading2"], fontName=font_name, fontSize=14)
    normal_style = ParagraphStyle("ChNormal", parent=styles["Normal"], fontName=font_name, fontSize=10, leading=16)
    cell_style = ParagraphStyle("ChCell", parent=styles["Normal"], fontName=font_name, fontSize=9, leading=13)

    # reportlab 的 Flowable 是动态类型，story 元素均为其子类（Paragraph/Spacer/Table）
    from typing import Any

    story: list[Any] = []
    story.append(Paragraph("fuscan 扫描报告", title_style))
    story.append(Spacer(1, 8 * mm))

    # 统计信息
    story.append(Paragraph("扫描统计", heading_style))
    story.append(Spacer(1, 3 * mm))
    stat_lines = [
        f"扫描路径: {report.root}",
        f"扫描时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"总文件数: {report.stats.total_files}",
        f"已扫描: {report.stats.scanned_files}",
        f"命中文件: {report.stats.matched_files}",
        f"总匹配数: {report.stats.total_matches}",
        f"跳过: {report.stats.skipped_files}",
        f"错误: {report.stats.errors}",
        f"耗时: {report.stats.duration_seconds:.2f}s",
    ]
    if report.cancelled:
        stat_lines.append("状态: 已取消")
    for line in stat_lines:
        story.append(Paragraph(line, normal_style))
    story.append(Spacer(1, 6 * mm))

    # 命中文件表格
    if not report.hits:
        story.append(Paragraph("未发现命中项。", normal_style))
    else:
        story.append(Paragraph(f"命中文件 ({len(report.hits)})", heading_style))
        story.append(Spacer(1, 3 * mm))
        # 表头
        header = [Paragraph(text, cell_style) for text in ["路径", "大小", "等级", "规则", "描述", "匹配数", "详情"]]
        # 表格行含 Paragraph 与字符串混合，用 list[Any] 容纳
        rows: list[list[Any]] = [header]
        for result in report.hits:
            try:
                rel = str(result.path.relative_to(report.root))
            except ValueError:
                rel = str(result.path)
            for hit in result.hits:
                rows.append(
                    [
                        Paragraph(rel, cell_style),
                        Paragraph(format_size(result.size), cell_style),
                        Paragraph(hit.severity.value, cell_style),
                        Paragraph(hit.rule_name, cell_style),
                        Paragraph(hit.match_description or "", cell_style),
                        Paragraph(str(hit.match_count), cell_style),
                        Paragraph(hit.detail, cell_style),
                    ]
                )
        col_widths = [50 * mm, 18 * mm, 14 * mm, 24 * mm, 24 * mm, 12 * mm, 32 * mm]
        table = Table(rows, colWidths=col_widths, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0887A0")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1DDE2")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F8FA")]),
                ]
            )
        )
        story.append(table)

    doc.build(story)
    return buf.getvalue()


def export_excel(report: ScanReport) -> bytes:
    """将扫描报告转换为 Excel 二进制数据（.xlsx）。

    使用 openpyxl 生成，含两个工作表：

    - "扫描汇总"：扫描路径、时间、统计信息
    - "命中明细"：每行一条规则命中（含描述列），严重等级着色
    """
    # 惰性导入：openpyxl 是核心依赖但导入较重
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Sheet1: 扫描汇总
    ws1 = wb.active
    assert ws1 is not None  # Workbook 新建时 active 一定非空，缩窄 pyrefly 类型推断
    ws1.title = "扫描汇总"
    bold_font = Font(bold=True)
    ws1["A1"] = "fuscan 扫描报告"
    ws1["A1"].font = Font(bold=True, size=14)
    summary_rows = [
        ("扫描路径", str(report.root)),
        ("扫描时间", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("总文件数", report.stats.total_files),
        ("已扫描", report.stats.scanned_files),
        ("命中文件", report.stats.matched_files),
        ("总匹配数", report.stats.total_matches),
        ("跳过", report.stats.skipped_files),
        ("错误", report.stats.errors),
        ("耗时(秒)", round(report.stats.duration_seconds, 2)),
        ("状态", "已取消" if report.cancelled else "完成"),
    ]
    for row_idx, (label, value) in enumerate(summary_rows, start=3):
        ws1.cell(row=row_idx, column=1, value=label).font = bold_font
        ws1.cell(row=row_idx, column=2, value=value)
    ws1.column_dimensions["A"].width = 16
    ws1.column_dimensions["B"].width = 50

    # Sheet2: 命中明细
    ws2 = wb.create_sheet("命中明细")
    headers = ["路径", "大小", "严重等级", "规则", "描述", "匹配数", "详情"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = bold_font
        cell.fill = PatternFill(start_color="0887A0", end_color="0887A0", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    row_idx = 2
    # 严重等级颜色映射
    severity_fills = {
        Severity.CRITICAL: PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid"),
        Severity.WARNING: PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid"),
        Severity.INFO: PatternFill(start_color="D4EFDF", end_color="D4EFDF", fill_type="solid"),
    }
    wrap_align = Alignment(wrap_text=True, vertical="top")
    for result in report.hits:
        try:
            rel = str(result.path.relative_to(report.root))
        except ValueError:
            rel = str(result.path)
        for hit in result.hits:
            ws2.cell(row=row_idx, column=1, value=rel).alignment = wrap_align
            ws2.cell(row=row_idx, column=2, value=format_size(result.size))
            severity_cell = ws2.cell(row=row_idx, column=3, value=hit.severity.value)
            severity_cell.fill = severity_fills.get(hit.severity, PatternFill())
            ws2.cell(row=row_idx, column=4, value=hit.rule_name).alignment = wrap_align
            ws2.cell(row=row_idx, column=5, value=hit.match_description).alignment = wrap_align
            ws2.cell(row=row_idx, column=6, value=hit.match_count)
            ws2.cell(row=row_idx, column=7, value=hit.detail).alignment = wrap_align
            row_idx += 1
    # 列宽
    col_widths = [50, 12, 12, 20, 20, 10, 40]
    for col_idx, width in enumerate(col_widths, start=1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = width
    # 冻结表头
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def save_report(report: ScanReport, path: Path) -> None:
    """将扫描报告保存到文件，根据扩展名自动选择格式。

    支持的扩展名：``.csv``/``.json``/``.txt``/``.pdf``/``.xlsx``。
    其他扩展名按文本格式输出。

    :param report: 扫描报告
    :param path: 目标文件路径；二进制格式（pdf/xlsx）写 bytes，文本格式写 UTF-8
    """
    ext = path.suffix.lower()
    if ext == ".pdf":
        path.write_bytes(export_pdf(report))
        return
    if ext == ".xlsx":
        path.write_bytes(export_excel(report))
        return
    content = report.to_format(ext.lstrip(".") if ext.lstrip(".") in ("csv", "json") else "text")
    path.write_text(content, encoding="utf-8")
